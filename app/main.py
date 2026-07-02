"""FastAPI application for Tekion report generation."""

from __future__ import annotations

import asyncio
import base64
import json
import re
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

from io import BytesIO

from app.aggregator import build_report
from app.config import load_config
from app.menu.admin_routes import router as menu_admin_router
from app.menu.routes import router as menu_router
from app.parsers.cdk.advisor_recap_csv import is_advisor_recap_csv
from app.parsers.cdk.export_pipeline import build_report_from_dynatron_exports
from app.parsers.cdk.opcode_analysis import is_opcode_analysis_report
from app.parsers.opcode import parse_opcode_report
from app.pdf_report import generate_pdf

app = FastAPI(title="Dealer Industries", version="0.2.0")
app.include_router(menu_router)
app.include_router(menu_admin_router)

STATIC_DIR = Path(__file__).resolve().parent / "static"
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


def _read_page(name: str) -> str:
    return (STATIC_DIR / name).read_text(encoding="utf-8")


@app.get("/favicon.ico", include_in_schema=False)
async def favicon() -> FileResponse:
    return FileResponse(STATIC_DIR / "logo.png", media_type="image/png")


@app.get("/", response_class=HTMLResponse)
async def home() -> str:
    return _read_page("reports.html")


@app.get("/reports", response_class=HTMLResponse)
async def reports() -> str:
    return _read_page("reports.html")


@app.get("/menu-builder", response_class=HTMLResponse)
async def menu_builder() -> str:
    return _read_page("menu-builder.html")


@app.get("/product-admin", response_class=HTMLResponse)
async def product_admin() -> str:
    return _read_page("product-admin.html")


@app.get("/menu-builder/settings", response_class=HTMLResponse)
async def menu_builder_settings() -> str:
    return _read_page("product-admin.html")


@app.get("/menu-builder/projects/{project_id}", response_class=HTMLResponse)
async def menu_builder_project(project_id: str) -> str:
    return _read_page("menu-project.html")


def _classify_excel(file_bytes: bytes) -> str:
    try:
        if is_opcode_analysis_report(file_bytes):
            return "cdk_opcode"
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        if "Summary" in sheets and "Customer Pay" in sheets:
            return "advisor"
        if "Dealership Pricing " in sheets or "Advisor Ops" in sheets:
            return "workflow"
    except Exception:
        pass
    return "unknown"


@app.post("/api/generate-cdk/stream")
async def generate_cdk_report_stream(
    photos: list[UploadFile] = File(..., description="All CDK report photos (recap + opcode pages)"),
    period_end: str = Form(default=""),
    store_name: str = Form(default=""),
    config_name: str = Form(default="classic_kia_carrollton"),
):
    if not photos:
        raise HTTPException(status_code=400, detail="Please upload at least one CDK report photo.")

    photo_bytes = [await f.read() for f in photos]
    loop = asyncio.get_running_loop()
    queue: asyncio.Queue = asyncio.Queue()

    def on_progress(event: dict) -> None:
        loop.call_soon_threadsafe(queue.put_nowait, event)

    async def run_job() -> None:
        from app.parsers.cdk.pipeline import build_report_from_cdk_photos

        try:
            report, pdf_bytes = await asyncio.to_thread(
                build_report_from_cdk_photos,
                photo_bytes,
                config_name=config_name.strip() or "classic_kia_carrollton",
                period_override=period_end.strip() or None,
                store_name_override=store_name.strip() or None,
                on_progress=on_progress,
            )
            safe_period = report.period_label.replace("/", "-")
            filename = f"advisor-report-{safe_period}.pdf"
            await queue.put(
                {
                    "type": "complete",
                    "percent": 100,
                    "message": "Report ready!",
                    "filename": filename,
                    "pdf_b64": base64.b64encode(pdf_bytes).decode("ascii"),
                    "advisors": len(report.advisors),
                    "total_kits": report.total_kits,
                }
            )
        except ValueError as exc:
            detail = str(exc)
            status = 503 if "GEMINI_API_KEY" in detail else 400
            await queue.put({"type": "error", "message": detail, "status": status})
        except Exception as exc:
            await queue.put(
                {
                    "type": "error",
                    "message": f"Failed to process CDK photos: {exc}",
                    "status": 500,
                }
            )
        finally:
            await queue.put(None)

    async def event_stream():
        task = asyncio.create_task(run_job())
        while True:
            item = await queue.get()
            if item is None:
                break
            yield f"data: {json.dumps(item)}\n\n"
        await task

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/generate-cdk")
async def generate_cdk_report(
    photos: list[UploadFile] = File(..., description="All CDK report photos (recap + opcode pages)"),
    period_end: str = Form(default=""),
    store_name: str = Form(default=""),
    config_name: str = Form(default="classic_kia_carrollton"),
):
    if not photos:
        raise HTTPException(status_code=400, detail="Please upload at least one CDK report photo.")

    photo_bytes = [await f.read() for f in photos]

    from app.parsers.cdk import build_report_from_cdk_photos

    try:
        report, pdf_bytes = build_report_from_cdk_photos(
            photo_bytes,
            config_name=config_name.strip() or "classic_kia_carrollton",
            period_override=period_end.strip() or None,
            store_name_override=store_name.strip() or None,
        )
    except ValueError as exc:
        detail = str(exc)
        if "GEMINI_API_KEY" in detail:
            raise HTTPException(status_code=503, detail=detail) from exc
        raise HTTPException(status_code=400, detail=detail) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to process CDK photos: {exc}") from exc

    safe_period = report.period_label.replace("/", "-")
    filename = f"advisor-report-{safe_period}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.post("/api/generate-dynatron-export")
async def generate_dynatron_export_report(
    opcode_file: UploadFile = File(..., description="Dynatron Op Code Analysis export (.xlsx)"),
    recap_file: UploadFile = File(..., description="Dynatron Service Advisor Recap export (.csv)"),
    period_end: str = Form(default=""),
    store_name: str = Form(default=""),
    config_name: str = Form(default="dynatron_bardahl"),
):
    opcode_bytes = await opcode_file.read()
    recap_bytes = await recap_file.read()

    if not is_opcode_analysis_report(opcode_bytes):
        raise HTTPException(
            status_code=400,
            detail="Could not identify the Op Code Analysis export. Please upload the Dynatron .xlsx report.",
        )
    if not is_advisor_recap_csv(recap_bytes):
        raise HTTPException(
            status_code=400,
            detail="Could not identify the Service Advisor Recap export. Please upload the Dynatron RAP .csv report.",
        )

    try:
        report, pdf_bytes = build_report_from_dynatron_exports(
            opcode_bytes,
            recap_bytes,
            config_name=config_name.strip() or "dynatron_bardahl",
            period_override=period_end.strip() or None,
            store_name_override=store_name.strip() or None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to generate Dynatron export report: {exc}") from exc

    safe_period = report.period_label.replace("/", "-")
    filename = f"advisor-report-{safe_period}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.post("/api/generate-cdk-export")
async def generate_cdk_export_report_alias(
    opcode_file: UploadFile = File(...),
    recap_file: UploadFile = File(...),
    period_end: str = Form(default=""),
    store_name: str = Form(default=""),
    config_name: str = Form(default="dynatron_bardahl"),
):
    """Backward-compatible alias for the Dynatron export endpoint."""
    return await generate_dynatron_export_report(
        opcode_file=opcode_file,
        recap_file=recap_file,
        period_end=period_end,
        store_name=store_name,
        config_name=config_name,
    )


@app.post("/api/generate")
async def generate_report(
    files: list[UploadFile] = File(..., description="Upload Tekion reports"),
    period_end: str = Form(default=""),
    store_name: str = Form(default=""),
):
    if not files or len(files) < 2:
        raise HTTPException(status_code=400, detail="Please upload the required Tekion reports.")

    advisor_bytes = None
    opcode_bytes = None

    for f in files:
        if not f.filename:
            continue
        ext = Path(f.filename).suffix.lower()
        content = await f.read()
        
        if ext == ".csv":
            opcode_bytes = content
        elif ext in {".xlsx", ".xlsm"}:
            cls = _classify_excel(content)
            if cls == "advisor":
                advisor_bytes = content

    if not advisor_bytes:
        raise HTTPException(
            status_code=400,
            detail="Could not identify the Advisor Performance Report. Please ensure you uploaded it."
        )
    if not opcode_bytes:
        raise HTTPException(
            status_code=400,
            detail="Could not identify the OP Code History Report. Please ensure you uploaded it."
        )

    # 1. Parse OP Code to get the Site Name
    try:
        opcode_result = parse_opcode_report(opcode_bytes)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse OP Code CSV: {exc}")

    site_name = store_name.strip() or opcode_result.site_name
    if not site_name or site_name.startswith("-1_"):
        site_name = store_name.strip() or "Unknown Dealership"

    # 2. Always load the universal default config
    try:
        config = load_config("default")
    except FileNotFoundError:
        raise HTTPException(status_code=500, detail="Universal configuration not found.")

    period_override = period_end.strip() or None

    try:
        report = build_report(
            advisor_source=advisor_bytes,
            opcode_source=opcode_bytes,
            config=config,
            period_override=period_override,
            store_name_override=site_name,
        )
        pdf_bytes = generate_pdf(report)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {exc}") from exc

    safe_period = report.period_label.replace("/", "-")
    filename = f"advisor-report-{safe_period}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/workflow-steps")
async def workflow_steps():
    from app.menu.template_data import WORKFLOW_STEPS
    return WORKFLOW_STEPS


@app.get("/health")
async def health():
    return {"status": "ok"}
