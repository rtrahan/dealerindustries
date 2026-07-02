"""Parse Tekion Advisor Performance Report (xlsx)."""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import BinaryIO

from openpyxl import load_workbook


@dataclass
class AdvisorMetrics:
    name: str
    ro_count: float
    bill_hrs: float
    elr: float
    parts_gp_pct: float
    hrs_per_ro: float


def _parse_money(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_pct(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_advisor_report(source: BinaryIO) -> dict[str, AdvisorMetrics]:
    """Return Customer Pay metrics keyed by advisor name."""
    if isinstance(source, bytes):
        source = BytesIO(source)
    wb = load_workbook(source, data_only=True, read_only=True)
    if "Summary" not in wb.sheetnames:
        raise ValueError("Advisor report missing 'Summary' sheet.")

    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Advisor Summary sheet is empty.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = {name: i for i, name in enumerate(headers)}

    required = ["Name", "Pay Type", "Bill Hrs", "RO Count", "ELR ($)", "Parts GP (%)"]
    for field in required:
        if field not in col:
            raise ValueError(f"Advisor Summary missing column: {field}")

    advisors: dict[str, AdvisorMetrics] = {}
    for row in rows[1:]:
        if not row or row[col["Name"]] is None:
            continue
        name = str(row[col["Name"]]).strip()
        if name.lower() == "total":
            continue
            
        pay_type = str(row[col["Pay Type"]]).strip()
        if pay_type != "Customer Pay":
            continue
        ro_count = float(row[col["RO Count"]] or 0)
        bill_hrs = float(row[col["Bill Hrs"]] or 0)
        elr = _parse_money(row[col["ELR ($)"]])
        parts_gp = _parse_pct(row[col["Parts GP (%)"]])
        hrs_per_ro = bill_hrs / ro_count if ro_count else 0.0
        advisors[name] = AdvisorMetrics(
            name=name,
            ro_count=ro_count,
            bill_hrs=bill_hrs,
            elr=elr,
            parts_gp_pct=parts_gp,
            hrs_per_ro=hrs_per_ro,
        )

    wb.close()
    if not advisors:
        raise ValueError("No Customer Pay advisor rows found in Summary sheet.")
    return advisors


def parse_period_from_advisor_filename(filename: str) -> str | None:
    """Try to extract a period end date from common filename patterns."""
    match = re.search(r"(\d{4})[-_]?(\d{2})[-_]?(\d{2})", filename)
    if match:
        y, m, d = match.groups()
        return f"{m}-{d}-{y}"
    return None
