"""Parse Dynatron Op Code Analysis Report Excel exports."""

from __future__ import annotations

import re
from io import BytesIO
from typing import BinaryIO, Union

from openpyxl import load_workbook

from app.parsers.opcode import OpCodeParseResult, OpCodeRow

Source = Union[bytes, BinaryIO]

_ADVISOR_RE = re.compile(
    r"^(?P<last>[^,]+),\s*(?P<first>[^\(]+)\s*\((?P<sa_no>\d+)\)$"
)
_PERIOD_RE = re.compile(
    r"From\s+(\d{2}-\d{2}-\d{4})\s+To\s+(\d{2}-\d{2}-\d{4})",
    re.IGNORECASE,
)


def _as_binary(source: Source) -> BinaryIO:
    if isinstance(source, bytes):
        return BytesIO(source)
    return source


def _parse_qty(value) -> int:
    if value is None:
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _find_data_sheet(wb) -> tuple[str, list[tuple]]:
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        if "Op Code" in header and "Advisor" in header and "# ROs" in header:
            return name, rows
    raise ValueError("Op Code Analysis export missing expected data sheet.")


def _parse_metadata(wb) -> tuple[str | None, str | None, str | None]:
    store_name = None
    period_start = None
    period_end = None

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=12, values_only=True))
        if not rows or str(rows[0][0] or "").strip() != "Report Details":
            continue
        for row in rows:
            if not row or row[0] is None:
                continue
            text = str(row[0]).strip()
            if text.startswith("Dealer:"):
                store_name = text.split(":", 1)[1].strip()
            elif text.startswith("From "):
                match = _PERIOD_RE.search(text)
                if match:
                    period_start, period_end = match.group(1), match.group(2)
        break

    return store_name, period_start, period_end


def parse_opcode_analysis_report(source: Source) -> OpCodeParseResult:
    """Parse Dynatron Op Code Analysis xlsx into kit sale rows."""
    stream = _as_binary(source)
    wb = load_workbook(stream, data_only=True, read_only=True)

    store_name, period_start, period_end = _parse_metadata(wb)
    _, rows = _find_data_sheet(wb)

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    opcode_rows: list[OpCodeRow] = []
    for row in rows[2:]:
        if not row:
            continue
        op_code = row[col["Op Code"]] if col.get("Op Code") is not None else None
        advisor = row[col["Advisor"]] if col.get("Advisor") is not None else None
        if op_code is None or advisor is None:
            continue

        op_code = str(op_code).strip()
        advisor_text = str(advisor).strip()
        if not op_code or op_code.lower() == "code":
            continue
        if "All Op Codes" in op_code or "All Advisors" in advisor_text:
            continue

        match = _ADVISOR_RE.match(advisor_text)
        if not match:
            continue

        qty = _parse_qty(row[col["# ROs"]])
        if qty <= 0:
            continue

        opcode_rows.append(
            OpCodeRow(
                part_number=op_code,
                sale_qty=qty,
                service_advisor=match.group("last").strip().upper(),
                pay_type="CEXP",
                source_code="",
            )
        )

    wb.close()

    if not opcode_rows:
        raise ValueError("No kit sale rows found in Op Code Analysis export.")

    return OpCodeParseResult(
        rows=opcode_rows,
        period_start=period_start,
        period_end=period_end,
        site_name=store_name,
    )


def is_opcode_analysis_report(source: Source) -> bool:
    """Return True when bytes look like a Dynatron Op Code Analysis export."""
    stream = _as_binary(source)
    try:
        wb = load_workbook(stream, data_only=True, read_only=True)
        _find_data_sheet(wb)
        wb.close()
        return True
    except Exception:
        return False
